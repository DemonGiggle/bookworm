from __future__ import annotations

from pathlib import Path, PurePosixPath
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..core.models import DocumentSection, EmbeddedImage, SourceDocument, SourceRef
from ..images.base import ImageAnalyzer
from .base import SourceAdapter
from .embedded_images import analyze_embedded_images, mime_type_for_filename


def _render_row(row: List[object]) -> str:
    return " | ".join("" if cell is None else str(cell) for cell in row).strip()


def _nearest_non_empty_row_text(rows_by_index, start_index: int, step: int, max_row: int) -> str:
    index = start_index + step
    while 1 <= index <= max_row:
        text = rows_by_index.get(index, "").strip()
        if text:
            return text
        index += step
    return ""


def _dedupe_non_empty(items: List[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for item in items:
        normalized = item.strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        result.append(normalized)
    return result


def _embedded_images_for_sheet(
    sheet,
    source_id: str,
    path: Path,
    rows_by_index,
    max_row: int,
    image_offset: int,
) -> List[EmbeddedImage]:
    embedded_images: List[EmbeddedImage] = []
    images = sorted(
        getattr(sheet, "_images", []),
        key=lambda image: (
            getattr(getattr(image, "anchor", None), "_from", None).row
            if getattr(getattr(image, "anchor", None), "_from", None) is not None
            else 0,
            getattr(getattr(image, "anchor", None), "_from", None).col
            if getattr(getattr(image, "anchor", None), "_from", None) is not None
            else 0,
        ),
    )
    for image in images:
        marker = getattr(getattr(image, "anchor", None), "_from", None)
        row_number = (marker.row + 1) if marker is not None else 1
        column_number = (marker.col + 1) if marker is not None else 1
        cell_ref = "{column}{row}".format(
            column=get_column_letter(column_number),
            row=row_number,
        )
        raw_path = str(getattr(image, "path", "")).strip()
        filename = PurePosixPath(raw_path).name
        if not filename:
            image_format = str(getattr(image, "format", "")).strip().lower() or "img"
            filename = "sheet-image-{index}.{extension}".format(
                index=image_offset,
                extension=image_format,
            )
        caption = rows_by_index.get(row_number, "")
        context_text = "\n".join(
            _dedupe_non_empty(
                [
                    _nearest_non_empty_row_text(rows_by_index, row_number, -1, max_row),
                    _nearest_non_empty_row_text(rows_by_index, row_number, 1, max_row),
                ]
            )
        )
        embedded_images.append(
            EmbeddedImage(
                image_id="{source_id}-image-{index}".format(
                    source_id=source_id,
                    index=image_offset,
                ),
                source_ref=SourceRef(
                    source_id=source_id,
                    source_path=str(path),
                    locator="embedded image {index} on sheet {sheet} near {cell}".format(
                        index=image_offset,
                        sheet=sheet.title,
                        cell=cell_ref,
                    ),
                ),
                filename=filename,
                mime_type=mime_type_for_filename(filename),
                data=image._data(),
                caption=caption,
                context_text=context_text,
            )
        )
        image_offset += 1
    return embedded_images


class SpreadsheetAdapter(SourceAdapter):
    supported_suffixes = (".xlsx", ".xlsm")
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def load(
        self,
        path: Path,
        image_analyzer: Optional[ImageAnalyzer] = None,
    ) -> SourceDocument:
        workbook = load_workbook(filename=str(path), data_only=True)
        source_id = path.stem.replace(" ", "-").lower()
        sections = []
        embedded_images: List[EmbeddedImage] = []
        image_offset = 1
        for sheet in workbook.worksheets:
            rows = []
            rows_by_index = {}
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                rendered = _render_row(list(row))
                if rendered:
                    rows.append(rendered)
                    rows_by_index[row_index] = rendered
            sections.append(
                DocumentSection(
                    heading=sheet.title,
                    content="\n".join(rows),
                    source_ref=SourceRef(
                        source_id=source_id,
                        source_path=str(path),
                        locator="sheet {title}".format(title=sheet.title),
                    ),
                )
            )
            sheet_images = _embedded_images_for_sheet(
                sheet=sheet,
                source_id=source_id,
                path=path,
                rows_by_index=rows_by_index,
                max_row=sheet.max_row,
                image_offset=image_offset,
            )
            embedded_images.extend(sheet_images)
            image_offset += len(sheet_images)
        notes, warnings = analyze_embedded_images(
            sections=sections,
            embedded_images=embedded_images,
            image_analyzer=image_analyzer,
        )
        return SourceDocument(
            source_id=source_id,
            path=path,
            media_type=self.media_type,
            title=path.stem,
            sections=sections,
            embedded_images=embedded_images,
            extraction_notes=notes,
            extraction_warnings=warnings,
        )
